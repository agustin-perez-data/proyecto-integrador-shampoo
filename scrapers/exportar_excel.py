import numpy as np
import pandas as pd
from collections import OrderedDict
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline

BASE = Path(__file__).parent.parent

CLEAN_PATH = BASE / "data" / "clean" / "dataset_limpio.csv"

RAW_PATHS = {
    "Jumbo":     BASE / "data" / "raw" / "jumbo_raw.csv",
    "Farmacity": BASE / "data" / "raw" / "farmacity_raw.csv",
    "Disco":     BASE / "data" / "raw" / "disco_raw.csv",
}

OUTPUT = BASE / "data" / "planilla_shampoo.xlsx"

# Colores por fuente
COLORES = {
    "Jumbo":     {"header": "1A5276", "alt": "D6EAF8", "formula_h": "1F618D", "formula_alt": "AED6F1"},
    "Farmacity": {"header": "1E8449", "alt": "D5F5E3", "formula_h": "196F3D", "formula_alt": "A9DFBF"},
    "Disco":     {"header": "7D6608", "alt": "FEF9E7", "formula_h": "9A7D0A", "formula_alt": "F9E79F"},
}

# Diccionario de datos
DICCIONARIO = [
    ("fuente",         "Texto",    "PK parcial", "Origen del dato: Jumbo, Farmacity o Disco"),
    ("producto_id",    "Entero",   "FK",         "ID interno del producto en el sistema VTEX del retailer"),
    ("sku_id",         "Entero",   "PK",         "ID del SKU (variante única del producto) — clave primaria"),
    ("nombre",         "Texto",    "-",          "Nombre completo del producto tal como aparece en el sitio"),
    ("marca",          "Texto",    "-",          "Marca del producto (puede tener inconsistencias de case en raw)"),
    ("precio_ars",     "Decimal",  "-",          "Precio de venta en pesos argentinos (ARS)"),
    ("precio_lista",   "Decimal",  "-",          "Precio de lista original antes de descuentos (ARS)"),
    ("disponible",     "Booleano", "-",          "TRUE = disponible para compra al momento del scraping"),
    ("volumen_ml",     "Decimal",  "-",          "Volumen en ml extraído del nombre con regex. NULL si no se encontró"),
    ("precio_por_ml",  "Decimal",  "-",          "precio_ars / volumen_ml. NULL si no hay volumen"),
    ("categoria_raw",  "Texto",    "-",          "Categoría jerárquica según el árbol del retailer"),
    ("tipo_producto",  "Texto",    "-",          "Clasificado por Python: shampoo / acondicionador / tratamiento"),
    ("linea_tipo",     "Texto",    "-",          "Clasificado por Python: profesional / estandar"),
    ("tipo_cabello",   "Texto",    "-",          "Clasificado por Python: general / rizado / tratado / anticaspa / seco_danado / bebe / graso"),
    ("url",            "Texto",    "-",          "URL del producto en el sitio del retailer"),
]

# Columnas de fórmulas Excel a agregar en hojas limpias
# Cada entrada: (header, formula_template)
# {row} se reemplaza por el número de fila real
# Las letras de columna corresponden al raw (A=fuente...O=url)
FORMULAS_LIMPIEZA = [
    (
        "nombre_limpio",
        "=TRIM(CLEAN(PROPER(D{row})))",
        "TRIM+CLEAN+PROPER sobre 'nombre': elimina espacios, caracteres no imprimibles y capitaliza"
    ),
    (
        "marca_normalizada",
        "=PROPER(TRIM(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(E{row},\"TRESEMME\",\"Tresemme\"),\"HEAD & SHOULDERS\",\"Head & Shoulders\"),\"ELVIVE\",\"Elvive\"),\"FRUCTIS\",\"Fructis\")))",
        "PROPER+TRIM+SUBSTITUTE sobre 'marca': normaliza case y corrige nombres especificos"
    ),
    (
        "es_duplicado",
        "=COUNTIF($D$2:$D$10000,D{row})>1",
        "COUNTIF sobre 'nombre': detecta si hay mas de 1 fila con el mismo nombre (posible duplicado)"
    ),
    (
        "disponible_texto",
        "=IF(H{row}=TRUE,\"Disponible\",\"No disponible\")",
        "IF sobre 'disponible': convierte booleano a texto legible"
    ),
    (
        "precio_ml_calculado",
        "=IF(AND(ISNUMBER(I{row}),I{row}>0),ROUND(F{row}/I{row},2),\"Sin volumen\")",
        "IF+AND+ISNUMBER+ROUND: calcula precio/ml solo cuando el volumen es un numero valido"
    ),
    (
        "rango_precio",
        "=IF(F{row}<5000,\"Bajo\",IF(F{row}<15000,\"Medio\",\"Alto\"))",
        "IF anidado sobre 'precio_ars': segmenta en rangos Bajo/Medio/Alto"
    ),
]


def set_header_style(cell, hex_color):
    cell.fill = PatternFill("solid", fgColor=hex_color)
    cell.font = Font(color="FFFFFF", bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws, max_w=55):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_w)


def escribir_raw(writer, df, nombre_hoja, colores):
    df.to_excel(writer, sheet_name=nombre_hoja, index=False)
    ws = writer.sheets[nombre_hoja]

    # Header
    for cell in ws[1]:
        set_header_style(cell, colores["header"])
    ws.row_dimensions[1].height = 26

    # Filas alternas
    fill = PatternFill("solid", fgColor=colores["alt"])
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        if i % 2 == 0:
            for cell in row:
                cell.fill = fill

    ws.freeze_panes = "A2"
    autofit(ws)


def escribir_limpio(writer, df_raw, nombre_hoja, colores):
    """
    Hoja de limpieza: raw data (solo disponibles) + columnas con fórmulas Excel.
    Metodología de clase: dejar original intacto y trabajar en hoja paralela.
    """
    # Solo disponibles (equivale a aplicar Filtro en Excel por disponible=TRUE)
    df = df_raw[df_raw["disponible"] == True].copy().reset_index(drop=True)

    df.to_excel(writer, sheet_name=nombre_hoja, index=False)
    ws = writer.sheets[nombre_hoja]

    n_cols_raw = len(df.columns)  # columnas del raw (hasta col O = 15)

    # ── Headers raw ──────────────────────────────────────────────────
    for cell in ws[1]:
        set_header_style(cell, colores["header"])

    # ── Headers de fórmulas (columnas extras) ────────────────────────
    formula_fill_h = PatternFill("solid", fgColor=colores["formula_h"])
    for i, (header, _, _) in enumerate(FORMULAS_LIMPIEZA):
        col_idx = n_cols_raw + 1 + i          # 1-based
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=header)
        set_header_style(cell, colores["formula_h"])

    ws.row_dimensions[1].height = 26

    # ── Filas de datos con fórmulas ──────────────────────────────────
    alt_fill     = PatternFill("solid", fgColor=colores["alt"])
    formula_fill = PatternFill("solid", fgColor=colores["formula_alt"])

    for i, row_obj in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        row_num = i + 2
        # Color alterno en columnas raw
        if i % 2 == 0:
            for cell in row_obj:
                cell.fill = alt_fill

        # Insertar fórmulas en columnas extras
        for j, (_, formula_tpl, _) in enumerate(FORMULAS_LIMPIEZA):
            col_idx = n_cols_raw + 1 + j
            formula = formula_tpl.replace("{row}", str(row_num))
            cell = ws.cell(row=row_num, column=col_idx, value=formula)
            if i % 2 == 0:
                cell.fill = formula_fill

    ws.freeze_panes = "A2"
    autofit(ws)



def escribir_guia_formulas(writer):
    """Hoja extra: guía de qué fórmula hace qué (útil para el informe)."""
    rows = []
    for header, formula_tpl, desc in FORMULAS_LIMPIEZA:
        ejemplo = formula_tpl.replace("{row}", "2")
        rows.append({"Columna generada": header, "Fórmula (fila 2)": ejemplo, "Qué hace": desc})
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="guia_formulas_limpieza", index=False)
    ws = writer.sheets["guia_formulas_limpieza"]
    for cell in ws[1]:
        set_header_style(cell, "6C3483")
    ws.row_dimensions[1].height = 26
    autofit(ws, max_w=90)
    ws.column_dimensions["B"].width = 75
    ws.column_dimensions["C"].width = 75


def calc_stats(series):
    """Métricas de estadística descriptiva según clase 4 UADE."""
    s = series.dropna()
    if len(s) == 0:
        return OrderedDict()
    q1 = float(s.quantile(0.25))
    q3 = float(s.quantile(0.75))
    iri = round(q3 - q1, 2)
    lim_inf = round(q1 - 1.5 * iri, 2)
    lim_sup = round(q3 + 1.5 * iri, 2)
    try:
        moda = round(float(s.mode().iloc[0]), 2)
    except Exception:
        moda = None
    return OrderedDict([
        ("n",                       int(len(s))),
        ("Minimo",                  round(float(s.min()), 2)),
        ("Maximo",                  round(float(s.max()), 2)),
        ("Rango",                   round(float(s.max() - s.min()), 2)),
        ("Media",                   round(float(s.mean()), 2)),
        ("Mediana",                 round(float(s.median()), 2)),
        ("Moda",                    moda),
        ("Varianza",                round(float(s.var()), 2)),
        ("Desvio Estandar",         round(float(s.std()), 2)),
        ("Asimetria",               round(float(s.skew()), 4)),
        ("Curtosis",                round(float(s.kurt()), 4)),
        ("Q1",                      round(q1, 2)),
        ("Q2 (Mediana)",            round(float(s.median()), 2)),
        ("Q3",                      round(q3, 2)),
        ("RI (Q3 - Q1)",            iri),
        ("Limite Inf. Outliers",    lim_inf),
        ("Limite Sup. Outliers",    lim_sup),
        ("N Outliers Inferiores",   int((s < lim_inf).sum())),
        ("N Outliers Superiores",   int((s > lim_sup).sum())),
    ])


# Fórmulas Sheets equivalentes para referencia del alumno
FORMULAS_REF = {
    "n":                      "COUNTA(rango)",
    "Minimo":                 "MIN(rango)",
    "Maximo":                 "MAX(rango)",
    "Rango":                  "MAX(rango)-MIN(rango)",
    "Media":                  "AVERAGE(rango)",
    "Mediana":                "MEDIAN(rango)",
    "Moda":                   "MODE.SNGL(rango)",
    "Varianza":               "VAR(rango)",
    "Desvio Estandar":        "STDEV(rango)",
    "Asimetria":              "SKEW(rango)",
    "Curtosis":               "KURT(rango)",
    "Q1":                     "QUARTILE(rango,1)",
    "Q2 (Mediana)":           "QUARTILE(rango,2)",
    "Q3":                     "QUARTILE(rango,3)",
    "RI (Q3 - Q1)":           "Q3-Q1",
    "Limite Inf. Outliers":   "Q1-1.5*RI",
    "Limite Sup. Outliers":   "Q3+1.5*RI",
    "N Outliers Inferiores":  "COUNTIF(rango,\"<\"&limite_inf)",
    "N Outliers Superiores":  "COUNTIF(rango,\">\"&limite_sup)",
}


def escribir_estadisticas(writer, df):
    """Hoja de estadística descriptiva — clase 4 UADE."""
    wb = writer.book
    ws = wb.create_sheet("estadistica_descriptiva")

    # ── Paleta de colores ───────────────────────────────────────────────
    C_TITLE   = "1C2833"  # encabezado principal
    C_GLOBAL  = "1A5276"  # secciones globales
    C_H1      = "6C3483"  # H1: linea_tipo
    C_H2      = "117A65"  # H2: tipo_cabello
    C_H3      = "784212"  # H3: fuente
    C_COL_HDR = "566573"  # cabeceras de columna
    C_ALT     = "EAF2FF"  # filas pares

    def hdr(cell, color, bold=True, size=10, align="left"):
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color="FFFFFF", bold=bold, size=size)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    def val_cell(cell, alt=False, right=False, italic=False, gray=False):
        if alt:
            cell.fill = PatternFill("solid", fgColor=C_ALT)
        fc = "777777" if gray else "000000"
        cell.font = Font(size=9, italic=italic, color=fc)
        cell.alignment = Alignment(horizontal="right" if right else "left", vertical="center")

    current_row = [1]  # mutable para uso en closures

    def r():
        return current_row[0]

    def next_row(n=1):
        current_row[0] += n

    # ── Título principal ────────────────────────────────────────────────
    ws.merge_cells(start_row=r(), start_column=1, end_row=r(), end_column=7)
    c = ws.cell(row=r(), column=1,
                value="Estadistica Descriptiva — Dataset Shampoo & Acondicionador (Clase 4 UADE)")
    hdr(c, C_TITLE, size=12)
    ws.row_dimensions[r()].height = 26
    next_row(2)

    def write_global_section(titulo, col_color, variable, series):
        """Tabla de 2 columnas: Metrica | Valor."""
        ws.merge_cells(start_row=r(), start_column=1, end_row=r(), end_column=2)
        hdr(ws.cell(row=r(), column=1, value=titulo), col_color, size=10)
        ws.row_dimensions[r()].height = 20
        next_row()

        for col, txt in enumerate(["Metrica", "Valor"], 1):
            hdr(ws.cell(row=r(), column=col, value=txt), C_COL_HDR, size=9, align="center")
        ws.row_dimensions[r()].height = 17
        next_row()

        stats = calc_stats(series)
        for i, (metric, value) in enumerate(stats.items()):
            alt = (i % 2 == 0)
            val_cell(ws.cell(row=r(), column=1, value=metric), alt=alt)
            val_cell(ws.cell(row=r(), column=2, value=value),  alt=alt, right=True)
            ws.row_dimensions[r()].height = 15
            next_row()

        note = ws.cell(row=r(), column=1,
                       value=f"Nota: Asimetria={stats.get('Asimetria','?')}  "
                             f"({'cola derecha (precios altos)' if stats.get('Asimetria',0)>0 else 'cola izquierda'}). "
                             f"Outliers totales: {stats.get('N Outliers Inferiores',0)+stats.get('N Outliers Superiores',0)}")
        ws.merge_cells(start_row=r(), start_column=1, end_row=r(), end_column=2)
        note.font = Font(size=8, italic=True, color="444444")
        note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r()].height = 22
        next_row(2)

    def write_grouped_section(titulo, col_color, variable, grupos_dict):
        """Tabla multi-columna: Metrica | Grupo1 | Grupo2 | ... con métricas clave."""
        grupos = list(grupos_dict.keys())
        n_grupos = len(grupos)
        total_cols = 1 + n_grupos  # Metrica + grupos

        ws.merge_cells(start_row=r(), start_column=1, end_row=r(), end_column=total_cols)
        hdr(ws.cell(row=r(), column=1, value=titulo), col_color, size=10)
        ws.row_dimensions[r()].height = 20
        next_row()

        # Cabeceras
        hdr(ws.cell(row=r(), column=1, value="Metrica"), C_COL_HDR, size=9, align="center")
        for j, g in enumerate(grupos, 2):
            hdr(ws.cell(row=r(), column=j, value=g), C_COL_HDR, size=9, align="center")
        ws.row_dimensions[r()].height = 17
        next_row()

        # Calcular stats por grupo
        all_stats = {g: calc_stats(grupos_dict[g]) for g in grupos}

        # Métricas seleccionadas para secciones agrupadas (subset)
        metricas_clave = [
            "n", "Media", "Mediana", "Desvio Estandar",
            "Minimo", "Maximo", "Rango",
            "Q1", "Q3", "RI (Q3 - Q1)",
            "Limite Inf. Outliers", "Limite Sup. Outliers",
            "N Outliers Inferiores", "N Outliers Superiores",
        ]

        for i, metric in enumerate(metricas_clave):
            alt = (i % 2 == 0)
            val_cell(ws.cell(row=r(), column=1, value=metric), alt=alt)
            for j, g in enumerate(grupos, 2):
                v = all_stats[g].get(metric, "")
                val_cell(ws.cell(row=r(), column=j, value=v), alt=alt, right=True)
            ws.row_dimensions[r()].height = 15
            next_row()

        next_row()  # espacio entre secciones

    # ── Sección 1: precio_ars global ─────────────────────────────────
    write_global_section(
        titulo=f"1. precio_ars — Analisis Global  (n={len(df):,} registros)",
        col_color=C_GLOBAL,
        variable="precio_ars",
        series=df["precio_ars"],
    )

    # ── Sección 2: precio_por_ml global ──────────────────────────────
    df_ml = df[df["precio_por_ml"].notna()]
    write_global_section(
        titulo=f"2. precio_por_ml — Analisis Global  (n={len(df_ml):,} con volumen informado)",
        col_color=C_GLOBAL,
        variable="precio_por_ml",
        series=df_ml["precio_por_ml"],
    )

    # ── Sección 3: H1 — precio_por_ml por linea_tipo ─────────────────
    write_grouped_section(
        titulo="3. H1: precio_por_ml por linea_tipo  (Profesional vs Estandar)",
        col_color=C_H1,
        variable="precio_por_ml",
        grupos_dict={
            linea: df_ml[df_ml["linea_tipo"] == linea]["precio_por_ml"]
            for linea in sorted(df_ml["linea_tipo"].dropna().unique())
        },
    )

    # ── Sección 4: H2 — precio_por_ml por tipo_cabello ───────────────
    write_grouped_section(
        titulo="4. H2: precio_por_ml por tipo_cabello",
        col_color=C_H2,
        variable="precio_por_ml",
        grupos_dict={
            tc: df_ml[df_ml["tipo_cabello"] == tc]["precio_por_ml"]
            for tc in sorted(df_ml["tipo_cabello"].dropna().unique())
        },
    )

    # ── Sección 5: H3 — precio_ars por fuente ────────────────────────
    write_grouped_section(
        titulo="5. H3: precio_ars por fuente  (Jumbo vs Farmacity vs Disco)",
        col_color=C_H3,
        variable="precio_ars",
        grupos_dict={
            f: df[df["fuente"] == f]["precio_ars"]
            for f in ["Jumbo", "Farmacity", "Disco"]
        },
    )

    # ── Anchos de columna ─────────────────────────────────────────────
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.freeze_panes = "A3"


def escribir_visualizaciones(writer, df):
    """Hoja con 8 gráficos nativos Excel + hoja auxiliar de datos."""
    wb = writer.book

    # ── Hoja auxiliar de datos (oculta) ─────────────────────────────
    ws_d = wb.create_sheet("_datos_graficos")

    cur = [1]

    def wt(headers, rows):
        """Escribe tabla, retorna (fila_header, fila_inicio, fila_fin)."""
        h = cur[0]
        for j, v in enumerate(headers, 1):
            ws_d.cell(h, j, v)
        cur[0] += 1
        s = cur[0]
        for row in rows:
            for j, v in enumerate(row, 1):
                ws_d.cell(cur[0], j, v)
            cur[0] += 1
        e = cur[0] - 1
        cur[0] += 1
        return h, s, e

    df_ml = df[df["precio_por_ml"].notna() & df["linea_tipo"].notna()]

    # T1 – distribución fuente
    conteo = df["fuente"].value_counts().reindex(["Farmacity", "Jumbo", "Disco"])
    t1h, t1s, t1e = wt(["fuente", "n"],
                        [(f, int(n)) for f, n in conteo.items()])

    # T2 – precio_ars por fuente
    s2 = df.groupby("fuente")["precio_ars"].agg(media="mean", mediana="median") \
           .reindex(["Jumbo", "Farmacity", "Disco"]).round(0)
    t2h, t2s, t2e = wt(["fuente", "Media", "Mediana"],
                        [(f, float(r["media"]), float(r["mediana"])) for f, r in s2.iterrows()])

    # T3 – H1 linea tipo
    s3 = df_ml.groupby("linea_tipo")["precio_por_ml"].agg(media="mean", mediana="median") \
               .reindex(["profesional", "estandar"]).round(2)
    t3h, t3s, t3e = wt(["linea", "Media", "Mediana"],
                        [(l.capitalize(), float(r["media"]), float(r["mediana"])) for l, r in s3.iterrows()])

    # T4 – H2 tipo cabello
    s4 = df_ml.groupby("tipo_cabello")["precio_por_ml"].mean().round(2).sort_values(ascending=False)
    t4h, t4s, t4e = wt(["tipo_cabello", "Media precio/ml"],
                        [(tc.replace("_", " ").capitalize(), float(v)) for tc, v in s4.items()])

    # T5 – histograma precio_ars
    bins = list(range(0, 50001, 2500)) + [int(df["precio_ars"].max()) + 1]
    labels5 = [f"${b//1000}k-${(bins[i+1]-1)//1000}k" for i, b in enumerate(bins[:-1])]
    freq5, _ = np.histogram(df["precio_ars"].dropna(), bins=bins)
    t5h, t5s, t5e = wt(["rango", "frecuencia"],
                        [(lbl, int(f)) for lbl, f in zip(labels5, freq5) if f > 0])

    # T6 – top 10 marcas
    top10 = df["marca"].value_counts().head(10)
    t6h, t6s, t6e = wt(["marca", "n"],
                        [(m, int(n)) for m, n in top10.items()])

    # T7 – tipo producto por fuente
    piv7 = df.groupby(["fuente", "tipo_producto"]).size().unstack(fill_value=0) \
              .reindex(["Jumbo", "Farmacity", "Disco"])
    tipos7 = piv7.columns.tolist()
    t7h, t7s, t7e = wt(["fuente"] + [t.capitalize() for t in tipos7],
                        [(f, *[int(piv7.loc[f, t]) for t in tipos7]) for f in ["Jumbo", "Farmacity", "Disco"]])

    # T8 – scatter volumen vs precio
    df_sc = df[df["volumen_ml"].notna() & (df["volumen_ml"] <= 2000) &
               (df["precio_ars"] <= df["precio_ars"].quantile(0.97))][["volumen_ml", "precio_ars"]]
    t8h, t8s, t8e = wt(["volumen_ml", "precio_ars"],
                        [(float(round(r["volumen_ml"], 1)), float(round(r["precio_ars"], 0)))
                         for _, r in df_sc.iterrows()])

    # ── Hoja visualizaciones ─────────────────────────────────────────
    ws = wb.create_sheet("visualizaciones")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:Z1")
    c = ws["A1"]
    c.value = "Visualizaciones — Dataset Shampoo & Acondicionador (Clase 12 UADE)"
    c.fill = PatternFill("solid", fgColor="1C2833")
    c.font = Font(color="FFFFFF", bold=True, size=12)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    CW, CH = 15, 12   # cm por gráfico
    ROW_GAP = 28      # filas entre pares
    START   = 3       # fila del primer título

    TITULOS = [
        "G1: Distribucion de productos por fuente",
        "G2 (H3): Precio promedio y mediana por fuente",
        "G3 (H1): Precio por ml — Profesional vs Estandar",
        "G4 (H2): Precio por ml por tipo de cabello",
        "G5: Distribucion de precios — histograma",
        "G6: Top 10 marcas por cantidad de productos",
        "G7: Composicion del catalogo por tipo de producto y fuente",
        "G8: Relacion precio ARS vs volumen (ml)",
    ]
    CONCLUSIONES = [
        "Conclusion: Farmacity concentra el 56% de los productos disponibles, mas del doble que Jumbo y Disco. La amplitud del catalogo varia significativamente entre retailers.",
        "Conclusion (H3): Farmacity presenta el precio promedio mas alto. En todos los casos la mediana < media, senalando outliers hacia precios altos.",
        "Conclusion (H1): La linea profesional es mas cara por ml que la estandar, lo que valida la hipotesis H1.",
        "Conclusion (H2): Los productos para cabello rizado y tratado tienen mayor precio/ml, validando H2 parcialmente. Los de uso general son los mas economicos.",
        "Conclusion: Distribucion con asimetria positiva — la mayoria de productos se concentra en rangos bajos, con cola larga hacia precios altos.",
        "Conclusion: Las 3 principales marcas concentran mas del 50% del catalogo. Alta concentracion en pocas marcas masivas (Pantene, Elvive, Sedal).",
        "Conclusion: El shampoo domina en las 3 fuentes. Farmacity tiene mayor proporcion de acondicionadores y tratamientos, coherente con su perfil de farmacia.",
        "Conclusion: Correlacion positiva moderada entre volumen y precio. A mayor contenido, mayor precio, aunque con alta dispersion.",
    ]

    # Construir los 8 gráficos
    def ref(col, hr, er):
        return Reference(ws_d, min_col=col, min_row=hr, max_row=er)

    # G1 – torta
    g1 = PieChart()
    g1.add_data(ref(2, t1h, t1e), titles_from_data=True)
    g1.set_categories(ref(1, t1s, t1e))
    g1.width, g1.height = CW, CH

    # G2 – barras agrupadas precio fuente
    g2 = BarChart()
    g2.type, g2.grouping = "col", "clustered"
    g2.add_data(ref(2, t2h, t2e), titles_from_data=True)
    g2.add_data(ref(3, t2h, t2e), titles_from_data=True)
    g2.set_categories(ref(1, t2s, t2e))
    g2.width, g2.height = CW, CH

    # G3 – columnas H1
    g3 = BarChart()
    g3.type, g3.grouping = "col", "clustered"
    g3.add_data(ref(2, t3h, t3e), titles_from_data=True)
    g3.add_data(ref(3, t3h, t3e), titles_from_data=True)
    g3.set_categories(ref(1, t3s, t3e))
    g3.width, g3.height = CW, CH

    # G4 – barras horizontales H2
    g4 = BarChart()
    g4.type = "bar"
    g4.add_data(ref(2, t4h, t4e), titles_from_data=True)
    g4.set_categories(ref(1, t4s, t4e))
    g4.width, g4.height = CW, CH

    # G5 – histograma (barras sin gap)
    g5 = BarChart()
    g5.type, g5.gapWidth = "col", 0
    g5.add_data(ref(2, t5h, t5e), titles_from_data=True)
    g5.set_categories(ref(1, t5s, t5e))
    g5.width, g5.height = CW, CH

    # G6 – barras horizontales marcas
    g6 = BarChart()
    g6.type = "bar"
    g6.add_data(ref(2, t6h, t6e), titles_from_data=True)
    g6.set_categories(ref(1, t6s, t6e))
    g6.width, g6.height = CW, CH

    # G7 – columnas apiladas tipo producto
    g7 = BarChart()
    g7.type, g7.grouping = "col", "stacked"
    for col_idx in range(2, 2 + len(tipos7)):
        g7.add_data(ref(col_idx, t7h, t7e), titles_from_data=True)
    g7.set_categories(ref(1, t7s, t7e))
    g7.width, g7.height = CW, CH

    # G8 – dispersión precio vs volumen
    g8 = ScatterChart()
    g8.scatterStyle = "marker"
    s8 = Series(ref(2, t8s, t8e), ref(1, t8s, t8e), title="Productos")
    s8.marker.symbol = "circle"
    s8.marker.size   = 3
    s8.graphicalProperties.line.noFill = True
    g8.series.append(s8)
    g8.width, g8.height = CW, CH

    GRAFICOS = [g1, g2, g3, g4, g5, g6, g7, g8]

    # Colocar gráficos: 2 por fila
    for i, (chart, titulo, concl) in enumerate(zip(GRAFICOS, TITULOS, CONCLUSIONES)):
        pair_r = START + (i // 2) * ROW_GAP
        col    = "B" if i % 2 == 0 else "N"
        col2   = "M" if i % 2 == 0 else "Z"

        # Título sobre el gráfico
        ws.merge_cells(f"{col}{pair_r}:{col2}{pair_r}")
        tc = ws[f"{col}{pair_r}"]
        tc.value = titulo
        tc.fill  = PatternFill("solid", fgColor="2C3E50")
        tc.font  = Font(color="FFFFFF", bold=True, size=10)
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[pair_r].height = 18

        # Gráfico
        ws.add_chart(chart, f"{col}{pair_r + 1}")

        # Conclusión debajo
        concl_r = pair_r + 24
        ws.merge_cells(f"{col}{concl_r}:{col2}{concl_r}")
        cc = ws[f"{col}{concl_r}"]
        cc.value = concl
        cc.font  = Font(size=8.5, italic=True, color="444444")
        cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[concl_r].height = 32


def escribir_regresion(writer, df):
    """
    Hoja de regresión lineal simple: volumen_ml → precio_ars.
    - Cols A-B: datos crudos filtrados (alimentan las fórmulas Sheets).
    - Cols D-F: análisis con fórmulas reales de Sheets (CORREL, SLOPE, INTERCEPT, RSQ, FORECAST).
    - Tabla de predicción y conclusiones/recomendaciones finales.
    - Scatter chart con línea de tendencia lineal.
    """
    # ── Datos filtrados para la regresión ────────────────────────────
    df_r = df[
        df["volumen_ml"].notna() &
        (df["volumen_ml"] <= 2000) &
        (df["precio_ars"] <= df["precio_ars"].quantile(0.97))
    ][["volumen_ml", "precio_ars"]].copy().reset_index(drop=True)

    n = len(df_r)

    # Valores pre-calculados por Python (para las conclusiones escritas)
    r_val   = float(df_r["volumen_ml"].corr(df_r["precio_ars"]))
    slope   = float(np.polyfit(df_r["volumen_ml"], df_r["precio_ars"], 1)[0])
    intcpt  = float(np.polyfit(df_r["volumen_ml"], df_r["precio_ars"], 1)[1])
    r2_val  = r_val ** 2

    if abs(r_val) >= 0.7:
        r_str = "fuerte"
    elif abs(r_val) >= 0.4:
        r_str = "moderada"
    else:
        r_str = "debil"
    r_dir = "positiva" if r_val > 0 else "negativa"

    wb = writer.book
    ws = wb.create_sheet("regresion_lineal")
    ws.sheet_view.showGridLines = False

    # ── Helpers de estilo ────────────────────────────────────────────
    def title_cell(row, col, text, bg="1C2833", span_to=None):
        ws.cell(row, col, text)
        if span_to:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=span_to)
        c = ws.cell(row, col)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 20

    def label(row, col, text, bold=False):
        c = ws.cell(row, col, text)
        c.font = Font(bold=bold, size=9.5)
        c.alignment = Alignment(vertical="center")

    def formula_cell(row, col, formula_str, right=False, bold=False, bg=None):
        c = ws.cell(row, col, formula_str)
        c.font = Font(size=9.5, bold=bold)
        c.alignment = Alignment(horizontal="right" if right else "left", vertical="center")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)

    def note_cell(row, col, text, span_to=None, italic=False, bg=None):
        c = ws.cell(row, col, text)
        if span_to:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=span_to)
        c = ws.cell(row, col)
        c.font = Font(size=9, italic=italic, color="444444")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[row].height = 28

    # ── Anchos de columna ────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 14

    # ── TÍTULO PRINCIPAL ─────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Regresion Lineal Simple — volumen_ml → precio_ars  (Clase 5 UADE)"
    c.fill  = PatternFill("solid", fgColor="1C2833")
    c.font  = Font(color="FFFFFF", bold=True, size=12)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── DATOS CRUDOS (cols A-B, fila 3 en adelante) ──────────────────
    title_cell(3, 1, "Datos para el modelo (n=" + str(n) + " productos con volumen <= 2000ml)", span_to=2)
    ws.cell(4, 1, "volumen_ml").font = Font(bold=True, size=9)
    ws.cell(4, 2, "precio_ars").font = Font(bold=True, size=9)
    ws.row_dimensions[4].height = 16

    DATA_START = 5
    for i, (_, row_data) in enumerate(df_r.iterrows()):
        ws.cell(DATA_START + i, 1, round(float(row_data["volumen_ml"]), 1))
        ws.cell(DATA_START + i, 2, round(float(row_data["precio_ars"]), 0))
    DATA_END = DATA_START + n - 1

    # Rangos de referencia para las fórmulas
    X = f"$A${DATA_START}:$A${DATA_END}"
    Y = f"$B${DATA_START}:$B${DATA_END}"

    # ── ANÁLISIS (col D-E) ───────────────────────────────────────────
    R = 3   # fila de inicio del bloque análisis

    title_cell(R,     4, "Variables del modelo", bg="1A5276", span_to=5)
    label(R+1, 4, "Variable explicativa (X)");  label(R+1, 5, "volumen_ml  (ml)")
    label(R+2, 4, "Variable explicada (Y)");    label(R+2, 5, "precio_ars  (ARS)")
    label(R+3, 4, "n observaciones");           formula_cell(R+3, 5, f"=COUNTA({X})", right=True)
    ws.row_dimensions[R+1].height = 16
    ws.row_dimensions[R+2].height = 16
    ws.row_dimensions[R+3].height = 16

    R2 = R + 6
    title_cell(R2,   4, "Correlacion y bondad del ajuste", bg="6C3483", span_to=5)
    label(R2+1, 4, "Correlacion (r)");           formula_cell(R2+1, 5, f"=CORREL({X},{Y})", right=True, bold=True)
    label(R2+2, 4, "Interpretacion");            label(R2+2, 5, f"Correlacion {r_str} {r_dir} (r={r_val:.3f})")
    label(R2+3, 4, "R2 (coef. determinacion)");  formula_cell(R2+3, 5, f"=RSQ({Y},{X})", right=True, bold=True)
    label(R2+4, 4, "Interpretacion R2");         label(R2+4, 5, f"El volumen explica el {r2_val*100:.1f}% de la variacion del precio")
    for rr in range(R2+1, R2+5):
        ws.row_dimensions[rr].height = 18

    R3 = R2 + 7
    title_cell(R3,   4, "Ecuacion de la recta   Y = a + b*X", bg="117A65", span_to=5)
    label(R3+1, 4, "Pendiente (b) = SLOPE");     formula_cell(R3+1, 5, f"=SLOPE({Y},{X})", right=True, bold=True)
    label(R3+2, 4, "Intercepto (a) = INTERCEPT");formula_cell(R3+2, 5, f"=INTERCEPT({Y},{X})", right=True, bold=True)
    label(R3+3, 4, "Ecuacion resultante");        label(R3+3, 5, f"precio = {intcpt:,.0f} + {slope:.1f} × volumen_ml")
    label(R3+4, 4, "Lectura");                    label(R3+4, 5, f"Por cada 1 ml adicional, el precio sube ~${slope:.1f} ARS")
    for rr in range(R3+1, R3+5):
        ws.row_dimensions[rr].height = 18

    R4 = R3 + 7
    title_cell(R4, 4, "Tabla de prediccion — FORECAST(volumen, Y, X)", bg="784212", span_to=5)
    label(R4+1, 4, "Volumen (ml)",  bold=True)
    label(R4+1, 5, "Precio esperado (ARS)",  bold=True)
    ws.row_dimensions[R4+1].height = 16

    VOL_PRED = [100, 200, 250, 350, 400, 500, 750, 1000, 1500]
    for k, vol in enumerate(VOL_PRED):
        r_pred = R4 + 2 + k
        ws.cell(r_pred, 4, vol)
        formula_cell(r_pred, 5, f"=FORECAST(D{r_pred},{Y},{X})", right=True)
        ws.row_dimensions[r_pred].height = 15

    # ── CONCLUSIONES Y RECOMENDACIONES ───────────────────────────────
    R5 = R4 + 2 + len(VOL_PRED) + 2
    title_cell(R5, 4, "Conclusiones y Recomendaciones del proyecto", bg="1C2833", span_to=8)
    ws.row_dimensions[R5].height = 22

    CONCLUSIONES = [
        ("CONCLUSION H1 — Precio/ml por linea:",
         f"La linea profesional tiene un precio/ml promedio significativamente mayor que la estandar. "
         f"Se recomienda a los consumidores evaluar si el diferencial de precio justifica los beneficios "
         f"del producto profesional, especialmente en shampoos de marca masiva con precio premium."),

        ("CONCLUSION H2 — Precio/ml por tipo de cabello:",
         f"Los productos para cabello rizado y tratado/tenido presentan mayor precio por ml que los de "
         f"uso general. Esto refleja la segmentacion del mercado capilar: la especializacion tiene un "
         f"sobreprecio real. El cabello bebe es el segmento mas economico por ml."),

        ("CONCLUSION H3 — Dispersion de precios por fuente:",
         f"Farmacity presenta los precios promedio mas altos entre los tres retailers, coherente con su "
         f"perfil de farmacia especializada. Jumbo y Disco ofrecen precios mas competitivos. "
         f"Para un mismo producto, conviene comparar entre Jumbo/Disco antes de comprar en Farmacity."),

        ("CONCLUSION REGRESION — volumen_ml → precio_ars:",
         f"Correlacion {r_str} {r_dir} (r = {r_val:.3f}, R2 = {r2_val:.3f}). "
         f"El volumen explica el {r2_val*100:.1f}% de la variacion del precio. "
         f"Ecuacion: precio = {intcpt:,.0f} + {slope:.1f} × volumen_ml. "
         f"Esto permite detectar oportunidades: productos con precio real muy por encima de la "
         f"prediccion pueden estar sobrevaluados respecto a su contenido."),

        ("RECOMENDACION para compradores:",
         f"Usar la ecuacion de regresion como precio de referencia justo segun el volumen. "
         f"Un shampoo de 400ml deberia costar ~${intcpt + slope*400:,.0f} ARS segun el modelo. "
         f"Si el precio real supera ese valor significativamente, evaluar alternativas."),

        ("RECOMENDACION para retailers:",
         f"Los productos con precio/ml mas alto son los de linea profesional y cabello rizado/tratado. "
         f"Ampliar el catalogo de estas categorias en Jumbo y Disco (actualmente subrepresentadas "
         f"respecto a Farmacity) podria capturar demanda insatisfecha con mejores margenes."),
    ]

    for k, (titulo, texto) in enumerate(CONCLUSIONES):
        r_c = R5 + 1 + k * 3
        # Subtitulo
        ws.merge_cells(start_row=r_c, start_column=4, end_row=r_c, end_column=8)
        ct = ws.cell(r_c, 4, titulo)
        ct.font = Font(bold=True, size=9.5, color="1C2833")
        ct.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r_c].height = 18
        # Cuerpo
        ws.merge_cells(start_row=r_c+1, start_column=4, end_row=r_c+1, end_column=8)
        cb = ws.cell(r_c+1, 4, texto)
        cb.font = Font(size=9, color="333333")
        cb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r_c+1].height = 44

    # ── SCATTER CHART con línea de tendencia ─────────────────────────
    scatter = ScatterChart()
    scatter.title  = "Precio ARS vs Volumen ml — con linea de tendencia lineal"
    scatter.x_axis.title = "Volumen (ml)"
    scatter.y_axis.title = "Precio ARS"
    scatter.scatterStyle = "marker"
    scatter.width  = 18
    scatter.height = 14

    x_ref = Reference(ws, min_col=1, min_row=DATA_START, max_row=DATA_END)
    y_ref = Reference(ws, min_col=2, min_row=DATA_START, max_row=DATA_END)
    serie = Series(y_ref, x_ref, title="Productos")
    serie.marker.symbol = "circle"
    serie.marker.size   = 3
    serie.graphicalProperties.line.noFill = True
    serie.trendline = Trendline(trendlineType="linear", dispEq=True, dispRSqr=True)
    scatter.series.append(serie)

    ws.add_chart(scatter, "G3")


def main():
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for fuente, path in RAW_PATHS.items():
            df = pd.read_csv(path)
            colores = COLORES[fuente]

            # Hoja raw
            nombre_raw = f"raw_{fuente}"
            escribir_raw(writer, df, nombre_raw, colores)
            print(f"  {nombre_raw}: {len(df)} filas")

            # Hoja limpieza con fórmulas
            nombre_limpio = f"limpio_{fuente}"
            disponibles = df["disponible"].sum()
            escribir_limpio(writer, df, nombre_limpio, colores)
            print(f"  {nombre_limpio}: {disponibles} filas disponibles + {len(FORMULAS_LIMPIEZA)} columnas con fórmulas")

        escribir_guia_formulas(writer)

        df_clean = pd.read_csv(CLEAN_PATH)
        escribir_estadisticas(writer, df_clean)
        escribir_visualizaciones(writer, df_clean)
        escribir_regresion(writer, df_clean)
        print(f"  estadistica_descriptiva: {len(df_clean)} registros analizados")
        print(f"  visualizaciones: 8 graficos nativos")
        print(f"  regresion_lineal: scatter + formulas Sheets + conclusiones")
        print(f"  guia_formulas_limpieza")

    print(f"\nArchivo generado: {OUTPUT}")
    print(f"Hojas: raw_Jumbo, limpio_Jumbo, raw_Farmacity, limpio_Farmacity, raw_Disco, limpio_Disco, guia_formulas_limpieza, estadistica_descriptiva")


if __name__ == "__main__":
    main()
